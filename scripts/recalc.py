"""
使用 LibreOffice 无界面打开工作簿、全表重算并保存，再用 openpyxl 扫描常见 Excel 错误。

不依赖 Microsoft Excel；需本机已安装 LibreOffice，且 `soffice` 在 PATH 中（Windows 可为 soffice.exe）。

用法:
  python recalc.py <文件.xlsx> [超时秒数]
  python recalc.py --scan-only <文件.xlsx>    # 仅扫描错误与公式数量，不重算

输出: JSON（stdout）
"""

from __future__ import annotations

import argparse
import json
import os
import platform
import shutil
import subprocess
import sys
from pathlib import Path

# 支持 `python scripts/recalc.py`（将 scripts 加入路径）
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from office_env import get_soffice_env

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "需要安装 openpyxl: pip install openpyxl"}, ensure_ascii=False),
        file=sys.stderr,
    )
    sys.exit(2)

# StarBasic 宏：全表计算后保存并关闭（OOXML 模块格式）
MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = (
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)


def _soffice_binary() -> str | None:
    for name in ("soffice", "soffice.exe"):
        p = shutil.which(name)
        if p:
            return p
    return None


def _libreoffice_macro_dir() -> Path:
    system = platform.system()
    if system == "Darwin":
        return Path(os.path.expanduser("~/Library/Application Support/LibreOffice/4/user/basic/Standard"))
    if system == "Windows":
        base = os.environ.get("APPDATA") or os.path.expanduser("~")
        if "APPDATA" in os.environ:
            return Path(os.environ["APPDATA"]) / "LibreOffice" / "4" / "user" / "basic" / "Standard"
        return Path(base) / "AppData" / "Roaming" / "LibreOffice" / "4" / "user" / "basic" / "Standard"
    # Linux 及其他类 Unix
    return Path(os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard"))


def _ensure_macro_file() -> tuple[bool, str]:
    macro_dir = _libreoffice_macro_dir()
    macro_file = macro_dir / MACRO_FILENAME
    if macro_file.is_file() and "RecalculateAndSave" in macro_file.read_text(encoding="utf-8", errors="ignore"):
        return True, str(macro_file)

    soffice = _soffice_binary()
    if not soffice:
        return False, "未找到 soffice，请安装 LibreOffice 并加入 PATH"

    macro_dir.mkdir(parents=True, exist_ok=True)
    try:
        # 触发用户配置目录生成（部分系统首次需要）
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=30,
            env=dict(get_soffice_env()),
            check=False,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    try:
        macro_file.write_text(RECALCULATE_MACRO, encoding="utf-8")
    except OSError as e:
        return False, f"无法写入宏文件: {e}"
    return True, str(macro_file)


def scan_workbook(path: str | Path) -> dict:
    """不重算，仅统计公式数量并扫描单元格中的常见错误串。"""
    path = Path(path)
    if not path.is_file():
        return {"error": f"文件不存在: {path}"}

    error_details: dict[str, list[str]] = {e: [] for e in EXCEL_ERRORS}
    total_errors = 0

    wb_values = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb_values.sheetnames:
            ws = wb_values[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is not None and isinstance(v, str):
                        for err in EXCEL_ERRORS:
                            if err in v:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
    finally:
        wb_values.close()

    formula_count = 0
    wb_form = load_workbook(path, data_only=False, read_only=True)
    try:
        for sheet_name in wb_form.sheetnames:
            ws = wb_form[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is not None and isinstance(v, str) and v.startswith("="):
                        formula_count += 1
    finally:
        wb_form.close()

    summary: dict[str, dict] = {}
    for err, locs in error_details.items():
        if locs:
            summary[err] = {"count": len(locs), "locations": locs[:50]}

    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
        "error_summary": summary,
    }


def recalc(path: str | Path, timeout: float = 30.0) -> dict:
    path = Path(path).resolve()
    if not path.is_file():
        return {"error": f"文件不存在: {path}"}

    ok, detail = _ensure_macro_file()
    if not ok:
        return {"error": detail}

    soffice = _soffice_binary()
    if not soffice:
        return {"error": "未找到 soffice"}

    abs_path = str(path)
    macro_url = (
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application"
    )
    cmd: list[str] = [
        soffice,
        "--headless",
        "--norestore",
        macro_url,
        abs_path,
    ]

    env = dict(get_soffice_env())
    creationflags = 0
    if platform.system() == "Windows" and hasattr(subprocess, "CREATE_NO_WINDOW"):
        creationflags = subprocess.CREATE_NO_WINDOW  # type: ignore[attr-defined]

    # Linux: 优先用 timeout 命令；其他系统用 subprocess 自带 timeout
    use_timeout_wrapper = platform.system() == "Linux" and shutil.which("timeout")
    if use_timeout_wrapper:
        cmd = ["timeout", str(int(timeout))] + cmd
        run_timeout = None
    else:
        run_timeout = timeout

    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            env=env,
            timeout=run_timeout,
            creationflags=creationflags,
        )
    except subprocess.TimeoutExpired:
        return {"error": f"LibreOffice 执行超时（{timeout}s）"}
    except FileNotFoundError:
        return {"error": "无法启动 soffice"}

    # Linux 上外层 `timeout` 命令在超时时常返回 124
    if use_timeout_wrapper and proc.returncode == 124:
        return {"error": f"LibreOffice 执行超时（{timeout}s）"}

    if proc.returncode not in (0, 124):
        err = (proc.stderr or proc.stdout or "LibreOffice 执行失败").strip()
        if "Module1" in err or "RecalculateAndSave" in err:
            return {"error": "LibreOffice 宏未正确加载，请检查用户目录下 Standard/Module1.xba"}
        return {"error": err[:2000] if err else "LibreOffice 执行失败"}

    return scan_workbook(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel 公式重算（LibreOffice）与错误扫描")
    parser.add_argument("path", help=".xlsx 文件路径")
    parser.add_argument(
        "timeout",
        nargs="?",
        type=int,
        default=30,
        help="超时秒数（默认 30，仅对非 Linux-timeout 包装路径生效时由 subprocess 使用）",
    )
    parser.add_argument(
        "--scan-only",
        action="store_true",
        help="仅扫描公式与错误，不调用 LibreOffice",
    )
    args = parser.parse_args()

    if args.scan_only:
        out = scan_workbook(args.path)
    else:
        out = recalc(args.path, timeout=float(args.timeout))

    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
